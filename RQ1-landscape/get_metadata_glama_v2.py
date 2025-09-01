# 获取截止20250630新增的目录的内容
# 修改请求配置，增加反反爬机制
# 找github链接
from bs4 import BeautifulSoup
import re
import xlwt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import WebDriverException
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import random
import traceback
import psutil

findName = re.compile(r'<a[^>]*>([^<]+)</a>')  # 获取mcp server的名称
findHref = re.compile(r'<a[^>]*href="([^"]+)"[^>]*>')  # 获取mcp server的链接
findOverview = re.compile(r'<div class="jrPWok dAaUHp fzopXD gnlqYH hWhrno">(.*?)</div>')  # 获取mcp server的概览
findLang = re.compile(r'<div[^>]*>\s*<div[^>]*style="[^"]*">\s*</div>([^<]+)</div>')  # 获取mcp server的实现语言
findGit = re.compile(r'href="([^"]*github\.com/[^"]+)"') # 获取github链接

def configure_chrome_options():
    """配置Chrome浏览器选项"""
    options = webdriver.ChromeOptions()
    # 反自动化检测设置
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    # 性能优化设置
    options.add_argument('--disable-extensions')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    # 忽略证书错误
    options.add_argument('--ignore-certificate-errors')
    # 无头模式选项（根据需要启用）
    # options.add_argument('--headless')
    return options

def main():
    # https://glama.ai/mcp/servers?attributes=hosting%3Aremote-capable%2Clanguage%3Apowershell%2Ccategory%3Aversion-control
    # 从这个之后开始爬
    # 算了 直接爬好了 也重复不了多少
    # hosting_list = ['remote-capable', 'hybrid']
    # language_list = ['javascript', 'html', 'c#', 'vbscript', 'jupyter notebook', 'powershell']
    # category_list = [
    #     'browser-automation', 'cloud-platforms', 'communication', 
    #     'customer-data-platforms', 'databases', 'developer-tools', 'file-systems', 
    #     'knowledge-and-memory', 'location-services', 'marketing', 'monitoring', 'search', 
    #     'travel-and-transportation', 'version-control', 'virtualization', 'finance', 
    #     'research-and-data', 'social-media', 'os-automation', 'note-taking', 'cloud-storage', 
    #     'calendar-management', 'ecommerce-and-retail', 'health-and-wellness', 'education-and-learning-tools', 
    #     'entertainment-and-media', 'home-automation-and-iot', 'customer-support', 'legal-and-compliance',
    #     'language-translation', 'speech-processing', 'image-and-video-processing', 'security-and-iam', 
    #     'games-and-gamification'
    # ]
    # hosting_list = ['hybrid']
    # language_list = ['typescript', 'python', 'go']
    # category_list = [
    #     'browser-automation', 'cloud-platforms', 'communication', 
    #     'customer-data-platforms', 'databases', 'developer-tools', 'file-systems', 
    #     'knowledge-and-memory', 'location-services', 'marketing', 'monitoring', 'search', 
    #     'travel-and-transportation', 'version-control', 'virtualization', 'finance', 
    #     'research-and-data', 'social-media', 'os-automation', 'note-taking', 'cloud-storage', 
    #     'calendar-management', 'ecommerce-and-retail', 'health-and-wellness', 'education-and-learning-tools', 
    #     'entertainment-and-media', 'home-automation-and-iot', 'customer-support', 'legal-and-compliance',
    #     'language-translation', 'speech-processing', 'image-and-video-processing', 'security-and-iam', 
    #     'games-and-gamification'
    # ]
    hosting_list = ['local-only', 'remote-capable', 'hybrid']

    language_list = ['typescript', 'python', 'go', 'javascript', 'html', 'c#', 'vbscript', 'jupyter notebook', 'powershell']

    # category_list = ['browser-automation', 'cloud-platforms', 'communication', 
    #                 'customer-data-platforms', 'databases', 'developer-tools', 'file-systems', 
    #                 'knowledge-and-memory', 'location-services', 'marketing', 'monitoring', 'search', 
    #                 'travel-and-transportation', 'version-control', 'virtualization', 'finance', 
    #                 'research-and-data', 'social-media', 'os-automation', 'note-taking', 'cloud-storage', 
    #                 'calendar-management', 'ecommerce-and-retail', 'health-and-wellness', 'education-and-learning-tools', 
    #                 'entertainment-and-media', 'home-automation-and-iot', 'customer-support', 'legal-and-compliance',
    #                 'language-translation', 'speech-processing', 'image-and-video-processing', 'security-and-iam', 
    #                 'games-and-gamification']
    # https://glama.ai/mcp/servers?attributes=author%3Aclaimed
    # author_list = ['claimed', 'official', 'unknown']
    category_list = ['app-automation', 'rag-systems', 'code-execution','autonomous-agents',
                         'agent-orchestration', 'web-scraping', 'content-management-systems',
                         'project-management', 'documentation-access', 'api-testing']

    # hosting_list = ['local-only', 'remote-capable', 'hybrid']

    # language_list = ['typescript', 'python', 'go', 'javascript', 'html', 'c#', 'vbscript', 'jupyter notebook', 'powershell']

    # category_list = ['browser-automation', 'cloud-platforms', 'communication', 
    #                 'customer-data-platforms', 'databases', 'developer-tools', 'file-systems', 
    #                 'knowledge-and-memory', 'location-services', 'marketing', 'monitoring', 'search', 
    #                 'travel-and-transportation', 'version-control', 'virtualization', 'finance', 
    #                 'research-and-data', 'social-media', 'os-automation', 'note-taking', 'cloud-storage', 
    #                 'calendar-management', 'ecommerce-and-retail', 'health-and-wellness', 'education-and-learning-tools', 
    #                 'entertainment-and-media', 'home-automation-and-iot', 'customer-support', 'legal-and-compliance',
    #                 'language-translation', 'speech-processing', 'image-and-video-processing', 'security-and-iam', 
    #                 'games-and-gamification']
    
    count = 0
    browser_restart_count = 0
    MAX_REQUESTS_PER_BROWSER = 50  # 每50次请求重启一次浏览器
    
    try:
        # 初始化浏览器
        driver = webdriver.Chrome(options=configure_chrome_options())
        driver.set_page_load_timeout(30)  # 设置页面加载超时时间
        
        for hosting in hosting_list:
            for language in language_list:
                for category in category_list:
                    count += 1
                    browser_restart_count += 1
                    
                    # 每50次请求重启浏览器释放资源
                    if browser_restart_count >= MAX_REQUESTS_PER_BROWSER:
                        print(f"达到{browser_restart_count}次请求，重启浏览器释放资源...")
                        driver.quit()
                        driver = webdriver.Chrome(options=configure_chrome_options())
                        browser_restart_count = 0
                    
                    # 内存监控：内存使用超过85%时重启浏览器
                    if psutil.virtual_memory().percent > 85:
                        print("内存使用超过85%，重启浏览器释放内存...")
                        driver.quit()
                        driver = webdriver.Chrome(options=configure_chrome_options())
                        browser_restart_count = 0
                    
                    print(f"\n开始第 {count} 次爬取：")
                    url = f"https://glama.ai/mcp/servers?attributes=hosting%3A{hosting}%2Clanguage%3A{language}%2Ccategory%3A{category}"
                    print(url)
                    item_data = [hosting, language, category]
                    
                    # 随机延迟1.5-4秒，避免请求过于频繁
                    delay = random.uniform(1.5, 4.0)
                    print(f"随机延迟 {delay:.2f} 秒...")
                    time.sleep(delay)
                    
                    # 获取数据并保存
                    datalist = getData(url, item_data, driver)
                    savepath = "glama_250630_v7.xlsx"
                    saveDataAdd(datalist, savepath)
                    print(f"数据已成功追加到文件 {savepath}")
                    
                    # 清除cookies减少资源占用
                    driver.delete_all_cookies()
    
    except Exception as e:
        print(f"程序发生异常: {str(e)}")
        traceback.print_exc()
    finally:
        # 确保最终关闭浏览器
        if 'driver' in locals():
            driver.quit()
        print("爬取结束！")

def getData(url, item_data, driver):
    """获取页面数据，包含智能重试机制"""
    datalist = []
    max_retries = 3
    
    for attempt in range(max_retries):
        try:
            print(f"尝试 #{attempt+1} 访问页面...")
            driver.get(url)
            
            # 等待页面加载
            time.sleep(3)
            
            # 模拟点击"Load More"按钮
            while True:
                try:
                    load_more_button = driver.find_element(By.XPATH, '//button[text()="Load More"]')
                    load_more_button.click()
                    print("点击'Load More'按钮...")
                    time.sleep(2)  # 等待新内容加载
                except:
                    break  # 如果没有找到按钮，退出循环
            
            # 获取页面源码
            html = driver.page_source
            
            # 解析页面
            soup = BeautifulSoup(html, "html.parser")
            items = soup.find_all('article', class_="bHfsGq! eXSGnJ jDOzgL jsnJKs mBFIl fPSBzf bnYmbW diVeFv")
            
            cnt = 0
            print("********************************************************************************************************************发现数据********************************************************************************************************************")
            for item in items:
                data = []
                item = str(item)
                cnt += 1
                
                # 提取名称
                name = re.findall(findName, item)
                data.append(name[0] if name else "No name found")
                
                # 提取链接和GitHub链接
                href = re.findall(findHref, item)
                if href:
                    full_href = "https://glama.ai" + href[0]
                    data.append(full_href)
                    
                    # 从链接中提取GitHub信息
                    atIndex = full_href.find("@")
                    gitLink = "https://github.com/" + full_href[atIndex + 1:] if atIndex != -1 else "No GitHub link"
                    data.append(gitLink)
                else:
                    data.extend(["No link found", "No GitHub link"])
                
                # 添加分类信息
                data.append(item_data[2])  # category
                data.append(item_data[0])  # hosting location
                
                # 提取概览
                overview = re.findall(findOverview, item)
                data.append(overview[0] if overview else "No overview found")
                
                # 添加语言
                data.append(item_data[1])  # language
                
                datalist.append(data)
                print(f"已提取 {cnt} 条数据")
            
            return datalist
        
        except WebDriverException as e:
            print(f"尝试 {attempt+1}/{max_retries} 失败: {str(e)}")
            if "net::ERR_CONNECTION_RESET" in str(e) or "handshake failed" in str(e):
                # 指数退避等待 (5, 25, 125秒)
                wait_time = 5 ** (attempt + 1)
                print(f"连接错误，等待 {wait_time} 秒后重试...")
                time.sleep(wait_time)
            else:
                # 其他错误直接重试
                time.sleep(5)
            
            # 最后一次尝试仍然失败
            if attempt == max_retries - 1:
                print(f"页面 {url} 爬取失败，跳过...")
                return []
    
    return []  # 默认返回空列表

def saveDataAdd(datalist, savepath):
    """追加数据到Excel文件"""
    try:
        # 尝试加载现有的Excel文件
        book = load_workbook(savepath)
        sheet = book.active
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的工作簿
        book = Workbook()
        sheet = book.active
        # 定义列标题
        col = (
            "Server Name",
            "Link",
            "Github Link",
            "Category",
            "Location（local/remote）",
            "Description",
            "Language",
        )
        # 写入列名
        for i in range(len(col)):
            sheet.cell(row=1, column=i + 1, value=col[i])

    # 获取当前工作表的最大行数
    max_row = sheet.max_row

    # 写入数据
    for i in range(len(datalist)):
        data = datalist[i]
        for j in range(len(data)):
            sheet.cell(row=max_row + i + 1, column=j + 1, value=data[j])

    # 保存工作簿
    book.save(savepath)

if __name__ == "__main__":
    main()