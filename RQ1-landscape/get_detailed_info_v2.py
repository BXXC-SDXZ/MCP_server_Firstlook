# 从github链接中爬取开发者信息、编程语言和github star数量
from bs4 import BeautifulSoup
import re
import openpyxl
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl import Workbook
import time

# 定义正则表达式
findOverview = re.compile(r'<p dir="auto">(.*?)</p>', re.DOTALL)
findLang = re.compile(r'<h2 class="h4 mb-3">Languages</h2>.*?<span class="color-fg-default text-bold mr-1">(.*?)</span>', re.DOTALL)
findGitStar = re.compile(r'<span[^>]*id="repo-stars-counter-star"[^>]*>(\d+)</span>')

def main():
    # 1.获取github url
    file_path = r'*******************************'
    save_path = r'*****************************'
    git_links = get_git_links(file_path)  # 读取excel文件中的github链接

    # 2.开始爬取（根据github链接)
    # 每爬取100个github链接，保存一次数据
    count = 0
    datalist_part = []  # 用于临时存储数据
    for git_link in git_links:
        count += 1
        print("开始第", count, "次爬取：")
        datalist = getData(git_link)
        if datalist:  # 如果爬取到数据
            datalist_part.extend(datalist)  # 将数据添加到临时存储列表中
        print("爬取完毕！")

        # 每100个链接保存一次数据
        if count % 3 == 0:
            print("保存数据：")
            saveDataAdd(datalist_part, save_path)  # 保存数据
            datalist_part = []  # 清空临时存储列表

    # 爬取完成后，保存剩余的数据
    if datalist_part:
        print("保存最后一批数据，共", len(datalist_part), "条数据：")
        saveDataAdd(datalist_part, save_path)  # 保存剩余数据

def get_git_links(file_path):
    # 打开Excel文件并读取github链接
    wb = openpyxl.load_workbook(file_path)  # 读取Excel文件
    sheet = wb.active  # 获取活动工作表
    git_links = []

    # 遍历D列，提取所有的GitHub链接
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):  # 从第二行开始读取D列（github链接）
        link = row[0].value
        if link:  # 如果链接不为空
            git_links.append(link)  # 添加链接
    return git_links

def getData(url):
    print("开始爬取数据：")
    datalist = []
    driver = webdriver.Chrome()  # 使用Chrome浏览器
    driver.get(url)
    
    # 等待页面加载
    time.sleep(3)

    # 获取页面源码
    html = driver.page_source
    driver.quit()

    # 解析页面
    soup = BeautifulSoup(html, "html.parser")
    
    items = soup.find_all('div', class_="application-main")
    for item in items:
        data = []
        item = str(item)

        # github链接
        git_link = url
        data.append(git_link)   

        # developer
        parts = url.split("/")
        developer = parts[3]
        data.append(developer)

        # github star
        git_star = re.findall(findGitStar, item)
        data.append(git_star[0] if git_star else "No git star found")    

        # 编程语言
        lang = re.findall(findLang, item)
        data.append(lang[0] if lang else "No lang found")
        
        datalist.append(data)

    return datalist

def saveDataAdd(datalist, savepath):
    try:
        # 尝试加载现有的Excel文件
        book = load_workbook(savepath)
        sheet = book.active
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的工作簿
        book = Workbook()
        sheet = book.active
        col = (
            "git链接",
            "开发者",
            "github star数量",
            "编程语言"
        )
        # 写入列名
        for i in range(len(col)):
            sheet.cell(row=1, column=i + 1, value=col[i])

    # 获取当前工作表的最大行数
    max_row = sheet.max_row

    # 写入数据
    for i in range(len(datalist)):
        data = datalist[i]
        for j in range(len(data)):
            sheet.cell(row=max_row + i + 1, column=j + 1, value=data[j])

    # 保存工作簿
    book.save(savepath)
    print(f"数据已成功追加到文件 {savepath}")

if __name__ == "__main__":
    main()