# 先获取简略信息，然后再根据网站上的链接找github链接以及详细的describtion和tool list
# 本文件获取的是简略信息，无github链接
from bs4 import BeautifulSoup
import re
import xlwt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import WebDriverException
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
import time
import random
import traceback
import psutil

findName = re.compile(r'<a[^>]*>(.*?)</a>', re.DOTALL)
findHref = re.compile(r'<a[^>]*href="([^"]+)"[^>]*>')
findOverview = re.compile(r'<p[^>]*>(.*?)</p>')

def configure_chrome_options():
    """配置Chrome浏览器选项"""
    options = webdriver.ChromeOptions()
    # 反自动化检测设置
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    # 性能优化设置
    options.add_argument('--disable-extensions')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    # 忽略证书错误
    options.add_argument('--ignore-certificate-errors')
    # 无头模式选项（根据需要启用）
    # options.add_argument('--headless')
    return options

def main():
    # 定义文件路径和列名
    key_words_file_path = r"*******************************"
    column_name = "关键词"
    # 调用getKeyWords函数并传递参数
    key_words = getKeyWords(key_words_file_path, column_name)
    
    count = 0
    browser_restart_count = 0
    MAX_REQUESTS_PER_BROWSER = 50  # 每50次请求重启一次浏览器
    
    try:
        # 初始化浏览器
        driver = webdriver.Chrome(options=configure_chrome_options())
        driver.set_page_load_timeout(30)  # 设置页面加载超时时间
        
        for keyword in key_words:
            count += 1
            browser_restart_count += 1
            
            # 每50次请求重启浏览器释放资源
            if browser_restart_count >= MAX_REQUESTS_PER_BROWSER:
                print(f"达到{browser_restart_count}次请求，重启浏览器释放资源...")
                driver.quit()
                driver = webdriver.Chrome(options=configure_chrome_options())
                browser_restart_count = 0
                # 重新加载后需要跳过当前关键词
                continue
            
            # 内存监控：内存使用超过85%时重启浏览器
            if psutil.virtual_memory().percent > 85:
                print("内存使用超过85%，重启浏览器释放内存...")
                driver.quit()
                driver = webdriver.Chrome(options=configure_chrome_options())
                browser_restart_count = 0
                # 重新加载后需要跳过当前关键词
                continue
            
            print(f"\n开始第 {count} 次爬取：")
            url = f"https://mcprepository.com/search/{keyword}"
            print(url)
            
            # 随机延迟1.5-4秒，避免请求过于频繁
            delay = random.uniform(1.5, 4.0)
            print(f"随机延迟 {delay:.2f} 秒...")
            time.sleep(delay)
            
            # 获取数据并保存
            datalist = getData(url, keyword, driver)
            savepath = "official_brief_250715_v1.xlsx"
            if datalist:
                saveDataAdd(datalist, savepath)
                print(f"数据已成功追加到文件 {savepath}")
            else:
                print(f"未获取到数据，跳过保存")
            
            # 清除cookies减少资源占用
            driver.delete_all_cookies()
    
    except Exception as e:
        print(f"程序发生异常: {str(e)}")
        traceback.print_exc()
    finally:
        # 确保最终关闭浏览器
        if 'driver' in locals():
            driver.quit()
        print("爬取结束！")

def getKeyWords(file_path, column_name):
    """
    从指定的Excel文件中读取关键词。
    :param file_path: Excel文件的路径
    :param column_name: 包含关键词的列名
    :return: 关键词列表
    """
    try:
        df = pd.read_excel(file_path)
        # 检查列是否存在
        if column_name not in df.columns:
            print(f"列名 '{column_name}' 不存在于文件中。")
            return []
        
        # 读取关键词并去除空值
        key_words = df[column_name].dropna().tolist()
        print(f"成功从文件中读取到 {len(key_words)} 个关键词。")
        return key_words
    except Exception as e:
        print(f"读取文件时出错：{e}")
        return []  # 如果出错，返回空列表

def extract_name(html):
    match = findName.search(html)
    if match:
        # 提取匹配的内容
        content = match.group(1)
        # 去除嵌套的HTML标签
        clean_name = re.sub(r'<[^>]+>', '', content).strip()
        return clean_name
    return None

def getData(url, keyword, driver):
    """获取页面数据，包含智能重试机制"""
    datalist = []
    max_retries = 3
    
    for attempt in range(max_retries):
        try:
            print(f"尝试 #{attempt+1} 访问页面...")
            driver.get(url)
            
            # 等待页面加载
            time.sleep(3)
            
            # 获取页面源码
            html = driver.page_source
            
            # 解析页面
            soup = BeautifulSoup(html, "html.parser")
            items = soup.find_all('div', class_="server")
            
            cnt = 0
            for item in items:
                data = []
                item = str(item)
                cnt += 1
                
                # 提取名称
                name = extract_name(item)
                # print("name: ", name)
                data.append(name or "No name found")
                
                # 提取链接
                href = re.findall(findHref, item)
                if href:
                    full_href = "https://mcprepository.com" + href[0]
                    # print("href: ", full_href)
                    data.append(full_href)
                else:
                    # print("href: No link found")
                    data.append("No link found")
                
                # 添加关键词
                # print("key_word: ", keyword)
                data.append(keyword)
                
                # 提取简介
                overview = re.findall(findOverview, item)
                if overview:
                    # print("overview: ", overview[0])
                    data.append(overview[0])
                else:
                    # print("overview: No overview found")
                    data.append("No overview found")
                
                datalist.append(data)
                print(f"已提取 {cnt} 条数据")
            
            return datalist
        
        except WebDriverException as e:
            print(f"尝试 {attempt+1}/{max_retries} 失败: {str(e)}")
            if "net::ERR_CONNECTION_RESET" in str(e) or "handshake failed" in str(e):
                # 指数退避等待 (5, 25, 125秒)
                wait_time = 5 ** (attempt + 1)
                print(f"连接错误，等待 {wait_time} 秒后重试...")
                time.sleep(wait_time)
            else:
                # 其他错误直接重试
                time.sleep(5)
            
            # 最后一次尝试仍然失败
            if attempt == max_retries - 1:
                print(f"页面 {url} 爬取失败，跳过...")
                return []
    
    return []  # 默认返回空列表

# 使用openpyxl库来处理Excel文件，以追加的方式写入数据
def saveDataAdd(datalist, savepath):
    try:
        # 尝试加载现有的Excel文件
        book = load_workbook(savepath)
        sheet = book.active
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的工作簿
        book = Workbook()
        sheet = book.active
        col = (
            "Server Name", # Server 名称
            "Link",        # 链接
            "Key Word",    # key word
            "Description", # 描述
        )
        # 写入列名
        for i in range(len(col)):
            sheet.cell(row=1, column=i + 1, value=col[i])

    # 获取当前工作表的最大行数
    max_row = sheet.max_row

    # 写入数据
    for i in range(len(datalist)):
        data = datalist[i]
        for j in range(len(data)):
            sheet.cell(row=max_row + i + 1, column=j + 1, value=data[j])

    # 保存工作簿
    book.save(savepath)
    print(f"数据已成功追加到文件 {savepath}")

if __name__ == "__main__":
    main()